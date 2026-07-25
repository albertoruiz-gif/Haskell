# -*- coding: utf-8 -*-
"""
Version 6: igual que v5, pero normaliza el nombre de archivo con la MISMA
funcion slugify que usa el nombre de producto (antes solo se separaba por
guiones, y fallaba con archivos nombrados con espacios/mayusculas, como
"Supermascarilla acidificante matizadora 240 g.webp").
"""
import re
import shutil
import openpyxl
from openpyxl.styles import Font
import os

RUTA_XLSX = "/sessions/gracious-lucid-meitner/mnt/outputs/catalogo-haskell/productos_haskell_img.xlsx"
CARPETA_IMG = "/sessions/gracious-lucid-meitner/mnt/Co Work/Haskell/imagenes_principales"
ORIGEN = "/sessions/gracious-lucid-meitner/mnt/Co Work/Haskell/catalogo-haskell/productos_haskell.xlsx"

def slugify(texto):
    t = texto.lower()
    reemplazos = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n","ü":"u","!":"","ç":"c","ã":"a","õ":"o","â":"a","ê":"e","(":"","+":"", ")":""}
    for k, v in reemplazos.items():
        t = t.replace(k, v)
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t

def tokens_tamano(cantidad, unidad):
    variantes = []
    cant_str = f"{cantidad:g}"
    variantes.append(f"{cant_str}{unidad}")
    if unidad == "ml" and cantidad >= 1000 and cantidad % 1000 == 0:
        variantes.append(f"{int(cantidad/1000)}l")
    return variantes

def contiene_tamano(slug, variantes):
    for v in variantes:
        if re.search(r"(^|-)" + re.escape(v) + r"($|-)", slug):
            return True
    return False

LINEAS_GENERICAS = {"supermascaras"}

def tokens_requeridos_linea(linea, nombre_pt):
    slug_linea = slugify(linea)
    if slug_linea in LINEAS_GENERICAS:
        genericos = {"supermascara", "de", "e", "do", "da"}
        toks = [t for t in slugify(nombre_pt).split("-") if not re.match(r"^\d+(ml|g|l)$", t)]
        return [t for t in toks if t not in genericos]
    return slug_linea.split("-")

shutil.copy(ORIGEN, RUTA_XLSX)

archivos_raw = sorted(f for f in os.listdir(CARPETA_IMG) if f.lower().endswith(".webp"))
grupos = {}
archivo_original = {}
for f in archivos_raw:
    nombre_sin_ext = re.sub(r"\.webp$", "", f, flags=re.IGNORECASE)
    nombre_sin_sufijo = re.sub(r"_(\d+)$", "", nombre_sin_ext)
    base = slugify(nombre_sin_sufijo)  # <- ahora normalizado igual que nombre_pt
    grupos.setdefault(base, []).append(f)

slugs_disponibles = [s for s in grupos if "kit" not in s.split("-")]
slugs_tokens = {s: set(s.split("-")) for s in slugs_disponibles}

wb = openpyxl.load_workbook(RUTA_XLSX)
ws = wb["Productos"]
header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
col_nombre_pt = header["nombre_original_pt"]
col_nombre_es = header["nombre_es"]
col_slug = header["slug"]
col_linea = header["linea"]
col_cantidad = header["cantidad"]
col_unidad = header["unidad"]
col_img_principal = header["imagen_principal_local"]
col_img_adicionales = header["imagenes_adicionales"]

FONT_BODY = Font(name="Arial", size=10)

exactos = 0
por_tokens = 0
por_linea_unica = 0
sin_match = []
detalle_aprox = []

for fila in range(2, ws.max_row + 1):
    slug_fila = ws.cell(row=fila, column=col_slug).value
    nombre_pt = ws.cell(row=fila, column=col_nombre_pt).value
    nombre_es = ws.cell(row=fila, column=col_nombre_es).value
    linea = ws.cell(row=fila, column=col_linea).value
    cantidad = ws.cell(row=fila, column=col_cantidad).value
    unidad = ws.cell(row=fila, column=col_unidad).value
    slug_pt = slugify(nombre_pt) if nombre_pt else slug_fila
    slug_es = slugify(nombre_es) if nombre_es else slug_fila
    slug_pt_tokens = set(slug_pt.split("-"))
    variantes_tam = tokens_tamano(cantidad, unidad)

    match = None
    tipo_match = None

    if slug_pt in grupos and "kit" not in slug_pt.split("-"):
        match, tipo_match = slug_pt, "exacto"
    elif slug_fila in grupos and "kit" not in slug_fila.split("-"):
        match, tipo_match = slug_fila, "exacto"
    elif slug_es in grupos and "kit" not in slug_es.split("-"):
        match, tipo_match = slug_es, "exacto (nombre_es)"
    else:
        req_linea = set(tokens_requeridos_linea(linea, nombre_pt))
        candidatos = [
            s for s in slugs_disponibles
            if contiene_tamano(s, variantes_tam) and req_linea.issubset(slugs_tokens[s])
        ]
        if candidatos:
            candidatos.sort(key=lambda s: (-len(slug_pt_tokens & slugs_tokens[s]), len(s)))
            match = candidatos[0]
            tipo_match = "por_tokens"
        else:
            candidatos_sin_tam = [s for s in slugs_disponibles if req_linea.issubset(slugs_tokens[s])]
            if len(candidatos_sin_tam) == 1:
                match = candidatos_sin_tam[0]
                tipo_match = "linea_unica"

    if match:
        archivos_grupo = sorted(grupos[match])
        principal = archivos_grupo[0]
        secundarios = archivos_grupo[1:]
        ws.cell(row=fila, column=col_img_principal, value=principal).font = FONT_BODY
        if secundarios:
            ws.cell(row=fila, column=col_img_adicionales, value=" | ".join(secundarios)).font = FONT_BODY
        if tipo_match.startswith("exacto"):
            exactos += 1
        elif tipo_match == "por_tokens":
            por_tokens += 1
            detalle_aprox.append((nombre_es, match))
        else:
            por_linea_unica += 1
            detalle_aprox.append((nombre_es, match + "  [unico de la linea]"))
    else:
        sin_match.append((fila, nombre_es, slug_pt))

wb.save(RUTA_XLSX)

print("=== RESUMEN v6 ===")
print("Total filas de productos:", ws.max_row - 1)
print("Coincidencias exactas:", exactos)
print("Coincidencias por tokens:", por_tokens)
print("Coincidencias por linea unica:", por_linea_unica)
print("SIN imagen encontrada:", len(sin_match))
if sin_match:
    print()
    print("--- SIN imagen ---")
    for fila, nombre_es, slug_pt in sin_match:
        print(f"  fila {fila}: {nombre_es}  (slug: {slug_pt})")
