# -*- coding: utf-8 -*-
"""
Agrega al catalogo Haskell la conversion de precios de BRL a soles peruanos (PEN)
y el precio de venta al publico (PVP) en Peru:
    precio_pen = precio_brl * tipo_cambio
    pvp_peru   = precio_pen * (1 + margen)
Tipo de cambio y margen son PARAMETROS editables (hoja Parametros_PEN), y las
columnas nuevas usan formulas reales que los referencian -- si el usuario cambia
el tipo de cambio o el margen, todo el catalogo recalcula solo.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RUTA = "/sessions/gracious-lucid-meitner/mnt/outputs/catalogo-haskell/productos_haskell.xlsx"

TIPO_CAMBIO = 0.67   # soles peruanos (PEN) por 1 real brasileno (BRL)
MARGEN = 0.25        # 25% de incremento sobre el precio convertido para fijar el PVP

wb = openpyxl.load_workbook(RUTA)

FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FONT_INPUT = Font(name="Arial", size=11, bold=True, color="0000FF")
FILL_HEADER = PatternFill(start_color="1F4A2E", end_color="1F4A2E", fill_type="solid")
FILL_ASSUMPTION = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ------------------------------------------------------------------
# Hoja de parametros (tipo de cambio y margen), para que las formulas
# del catalogo los referencien en vez de tener el numero fijo repetido
# en cada fila.
# ------------------------------------------------------------------
if "Parametros_PEN" in wb.sheetnames:
    del wb["Parametros_PEN"]
wsp = wb.create_sheet("Parametros_PEN", 1)  # la coloca como 2da hoja
wsp["A1"] = "Parametro"
wsp["B1"] = "Valor"
wsp["A1"].font = FONT_HEADER
wsp["B1"].font = FONT_HEADER
wsp["A1"].fill = FILL_HEADER
wsp["B1"].fill = FILL_HEADER

wsp["A2"] = "Tipo de cambio (soles PEN por 1 real BRL)"
wsp["B2"] = TIPO_CAMBIO
wsp["B2"].font = FONT_INPUT
wsp["B2"].fill = FILL_ASSUMPTION
wsp["B2"].number_format = "0.00"

wsp["A3"] = "Margen sobre el precio convertido para fijar el PVP en Peru"
wsp["B3"] = MARGEN
wsp["B3"].font = FONT_INPUT
wsp["B3"].fill = FILL_ASSUMPTION
wsp["B3"].number_format = "0%"

wsp["A5"] = "Formula aplicada"
wsp["A5"].font = Font(name="Arial", size=10, bold=True)
wsp["A6"] = "precio_pen = precio_brl x tipo_cambio"
wsp["A7"] = "pvp_peru_pen = precio_pen x (1 + margen)"
wsp["A6"].font = FONT_BODY
wsp["A7"].font = FONT_BODY

wsp["A9"] = "Nota"
wsp["A9"].font = Font(name="Arial", size=10, bold=True)
wsp["A10"] = ("Celdas B2 y B3 son los unicos valores editables. El resto del catalogo "
              "(hoja Productos) usa formulas que apuntan a estas dos celdas: si se "
              "actualiza el tipo de cambio o el margen aca, todos los precios en soles "
              "y el PVP se recalculan automaticamente en toda la hoja Productos.")
wsp["A10"].font = FONT_BODY
wsp["A10"].alignment = Alignment(wrap_text=True, vertical="top")
wsp.row_dimensions[10].height = 45

wsp.column_dimensions["A"].width = 55
wsp.column_dimensions["B"].width = 14

# ------------------------------------------------------------------
# Columnas nuevas en la hoja Productos
# ------------------------------------------------------------------
ws = wb["Productos"]

header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
col_reg_brl = header["precio_regular_brl"]
col_ofe_brl = header["precio_oferta_brl"]
letra_reg_brl = get_column_letter(col_reg_brl)
letra_ofe_brl = get_column_letter(col_ofe_brl)

nuevas_columnas = [
    "precio_regular_pen",
    "pvp_peru_regular_pen",
    "precio_oferta_pen",
    "pvp_peru_oferta_pen",
]

col_inicio = ws.max_column + 1
for i, nombre in enumerate(nuevas_columnas):
    c = col_inicio + i
    cell = ws.cell(row=1, column=c, value=nombre)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(c)].width = 18

col_preg_pen, col_pvp_reg_pen, col_pofe_pen, col_pvp_ofe_pen = range(col_inicio, col_inicio + 4)
letra_preg_pen = get_column_letter(col_preg_pen)
letra_pvp_reg_pen = get_column_letter(col_pvp_reg_pen)
letra_pofe_pen = get_column_letter(col_pofe_pen)
letra_pvp_ofe_pen = get_column_letter(col_pvp_ofe_pen)

for fila in range(2, ws.max_row + 1):
    f_preg = f"=IF(ISBLANK({letra_reg_brl}{fila}),\"\",{letra_reg_brl}{fila}*Parametros_PEN!$B$2)"
    f_pvp_reg = f"=IF(ISBLANK({letra_reg_brl}{fila}),\"\",{letra_reg_brl}{fila}*Parametros_PEN!$B$2*(1+Parametros_PEN!$B$3))"
    f_pofe = f"=IF(ISBLANK({letra_ofe_brl}{fila}),\"\",{letra_ofe_brl}{fila}*Parametros_PEN!$B$2)"
    f_pvp_ofe = f"=IF(ISBLANK({letra_ofe_brl}{fila}),\"\",{letra_ofe_brl}{fila}*Parametros_PEN!$B$2*(1+Parametros_PEN!$B$3))"

    for col, formula in (
        (col_preg_pen, f_preg),
        (col_pvp_reg_pen, f_pvp_reg),
        (col_pofe_pen, f_pofe),
        (col_pvp_ofe_pen, f_pvp_ofe),
    ):
        cell = ws.cell(row=fila, column=col, value=formula)
        cell.font = FONT_BODY
        cell.number_format = '#,##0.00 "S/"'

ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"

wb.save(RUTA)
print("OK - columnas agregadas:", nuevas_columnas)
print("Hoja Parametros_PEN creada con tipo_cambio =", TIPO_CAMBIO, "y margen =", MARGEN)
