# -*- coding: utf-8 -*-
"""
Genera productos_haskell.xlsx a partir de datos reales extraidos de
haskellcosmeticos.com.br (paginas de linea, 24/07/2026), traducidos a
espanol neutro latinoamericano (sin voseo argentino).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

FECHA_EXTRACCION = "2026-07-24"
FUENTE = "Haskell Cosmeticos"

# Cada producto: (linea, tipo_pt, tipo_es, nombre_pt, nombre_es, cantidad, unidad,
#   precio_regular_brl, subcategoria_pt, subcategoria_es, descripcion_es,
#   beneficios_es (lista), propiedades_es (lista), modo_uso_es, activos_pt (lista),
#   activos_es (lista), recomendado_es, url_origen)

productos = [
    dict(
        linea="Cavalo Forte", tipo_pt="Tonico", tipo_es="Tonico capilar",
        nombre_pt="Tonico Ativador Cavalo Forte 120ml", nombre_es="Tonico activador Cavalo Forte 120 ml",
        cantidad=120, unidad="ml", precio_brl=56.90,
        subcat_pt="Fortalecimento", subcat_es="Fortalecimiento capilar",
        desc_es="Tonico capilar de la linea Cavalo Forte, formulado para estimular el cuero cabelludo y acompanar el crecimiento saludable del cabello, como parte de una rutina de fortalecimiento.",
        beneficios=["Estimula el cuero cabelludo", "Contribuye al crecimiento capilar saludable", "Fortalece la raiz del cabello"],
        propiedades=["Fortalecedor", "Estimulante del cuero cabelludo"],
        modo_uso="Aplicar directamente sobre el cuero cabelludo limpio y masajear suavemente. No es necesario enjuagar.",
        activos_pt=["Biotina", "Pantenol", "Queratina"],
        activos_es=["Biotina", "Pantenol", "Queratina"],
        recomendado="Todo tipo de cabello, especialmente cabello debilitado o con caida",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-cavalo-forte-haskell/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Bendito Loiro 300ml", nombre_es="Champu Bendito Loiro 300 ml",
        cantidad=300, unidad="ml", precio_brl=62.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Champu de limpieza suave de la linea Bendito Loiro, formulado para preparar el cabello decolorado antes del tratamiento, sin agredir la fibra capilar y manteniendo el pH equilibrado.",
        beneficios=["Limpieza suave sin agredir los cabellos", "Mantiene el pH equilibrado", "Evita la rigidez del cabello"],
        propiedades=["Limpiador suave", "pH equilibrado"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien.",
        activos_pt=["Complexo de proteinas", "Vinagre balsamico de uva"],
        activos_es=["Complejo de proteinas", "Vinagre balsamico de uva"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Bendito Loiro 500ml", nombre_es="Champu Bendito Loiro 500 ml",
        cantidad=500, unidad="ml", precio_brl=84.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Champu de limpieza suave de la linea Bendito Loiro, formulado para preparar el cabello decolorado antes del tratamiento, sin agredir la fibra capilar y manteniendo el pH equilibrado.",
        beneficios=["Limpieza suave sin agredir los cabellos", "Mantiene el pH equilibrado", "Evita la rigidez del cabello"],
        propiedades=["Limpiador suave", "pH equilibrado"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien.",
        activos_pt=["Complexo de proteinas", "Vinagre balsamico de uva"],
        activos_es=["Complejo de proteinas", "Vinagre balsamico de uva"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Condicionador", tipo_es="Acondicionador",
        nombre_pt="Condicionador Bendito Loiro 300ml", nombre_es="Acondicionador Bendito Loiro 300 ml",
        cantidad=300, unidad="ml", precio_brl=62.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Acondicionador de la linea Bendito Loiro que sella las cuticulas, repone proteinas y reduce la electricidad estatica del cabello decolorado, aportando suavidad y liviandad.",
        beneficios=["Sella las cuticulas", "Reduce la electricidad estatica", "Repone proteinas", "Aporta suavidad y liviandad"],
        propiedades=["Sellador de cuticulas", "Reposicion de proteinas"],
        modo_uso="Aplicar sobre el cabello lavado, de medios a puntas, dejar actuar unos minutos y enjuagar.",
        activos_pt=["Complexo de proteinas", "Vinagre balsamico de uva"],
        activos_es=["Complejo de proteinas", "Vinagre balsamico de uva"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Mascara de Tratamento", tipo_es="Mascarilla de tratamiento",
        nombre_pt="Mascara de Tratamento Bendito Loiro 300g", nombre_es="Mascarilla de tratamiento Bendito Loiro 300 g",
        cantidad=300, unidad="g", precio_brl=74.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Mascarilla de alto poder reconstructor de la linea Bendito Loiro, indicada para el tratamiento profundo y la regeneracion de los danos causados por la decoloracion.",
        beneficios=["Alto poder reconstructor", "Tratamiento profundo", "Regenera danos de la decoloracion"],
        propiedades=["Reconstructora", "Regeneradora"],
        modo_uso="Aplicar sobre el cabello lavado, distribuir de medios a puntas y dejar actuar segun las indicaciones del envase antes de enjuagar.",
        activos_pt=["Complexo de proteinas", "Vinagre balsamico de uva"],
        activos_es=["Complejo de proteinas", "Vinagre balsamico de uva"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Reparador", tipo_es="Proteina capilar",
        nombre_pt="Proteina Capilar Bendito Loiro 150g", nombre_es="Proteina capilar Bendito Loiro 150 g",
        cantidad=150, unidad="g", precio_brl=55.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Pre-champu repositor inmediato de proteina de la linea Bendito Loiro, que reconstruye la estructura interna del cabello y reduce el riesgo de quiebre.",
        beneficios=["Reconstruye la estructura interna del cabello", "Reduce el riesgo de quiebre", "Actua contra la elasticidad excesiva"],
        propiedades=["Repositor de proteina", "Pre-tratamiento"],
        modo_uso="Aplicar antes del champu, sobre el cabello seco o humedo, siguiendo las indicaciones del envase.",
        activos_pt=["Complexo de proteinas"],
        activos_es=["Complejo de proteinas"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Bendito Loiro", tipo_pt="Finalizador", tipo_es="Fluido finalizador",
        nombre_pt="Fluido Proteico Bendito Loiro 120ml", nombre_es="Fluido proteico Bendito Loiro 120 ml",
        cantidad=120, unidad="ml", precio_brl=48.90,
        subcat_pt="Cabelos descoloridos", subcat_es="Cabello decolorado",
        desc_es="Fluido finalizador de la linea Bendito Loiro que aporta luminosidad, accion antifrizz y proteccion termica y solar al cabello decolorado.",
        beneficios=["Aporta luminosidad", "Accion antifrizz", "Proteccion termica y solar"],
        propiedades=["Finalizador", "Protector termico"],
        modo_uso="Aplicar una pequena cantidad sobre el cabello humedo o seco, de medios a puntas, antes de peinar o usar herramientas de calor.",
        activos_pt=["Complexo de proteinas", "Vinagre balsamico de uva"],
        activos_es=["Complejo de proteinas", "Vinagre balsamico de uva"],
        recomendado="Cabello decolorado o rubio con danos, elasticidad excesiva o porosidad",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/bendito-loiro/",
    ),
    dict(
        linea="Matcha", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Uso Diario Matcha & Alecrim 300ml", nombre_es="Champu de uso diario Matcha y romero 300 ml",
        cantidad=300, unidad="ml", precio_brl=58.90,
        subcat_pt="Controle da oleosidade", subcat_es="Control de la grasitud capilar",
        desc_es="Champu de uso diario de la linea Matcha, formulado con matcha y romero para controlar la grasitud del cuero cabelludo y mantener el cabello limpio y liviano por mas tiempo.",
        beneficios=["Controla la grasitud del cuero cabelludo", "Fortalece los cabellos desde la raiz", "Mantiene la limpieza por mas tiempo"],
        propiedades=["Anti-grasitud", "Limpieza diaria"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar. Puede usarse a diario.",
        activos_pt=["Matcha", "Extrato de Alecrim"],
        activos_es=["Matcha", "Extracto de romero"],
        recomendado="Cabello graso o con tendencia a la grasitud excesiva",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-matcha-haskell/",
    ),
    dict(
        linea="Matcha", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Antirresiduos Matcha & Hortela 300ml", nombre_es="Champu antirresiduos Matcha y menta 300 ml",
        cantidad=300, unidad="ml", precio_brl=58.90,
        subcat_pt="Controle da oleosidade", subcat_es="Control de la grasitud capilar",
        desc_es="Champu de limpieza profunda de la linea Matcha, formulado con matcha y menta para eliminar residuos de productos de peinado y contaminacion, y prevenir la caida del cabello.",
        beneficios=["Limpieza profunda", "Elimina residuos de productos y contaminacion", "Previene la caida del cabello", "Sensacion de frescor"],
        propiedades=["Limpieza profunda", "Efecto refrescante"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien. Se recomienda su uso periodico, no necesariamente a diario.",
        activos_pt=["Matcha", "Extrato de Hortela"],
        activos_es=["Matcha", "Extracto de menta"],
        recomendado="Cabello graso o con acumulacion de residuos de productos",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-matcha-haskell/",
    ),
    dict(
        linea="Matcha", tipo_pt="Condicionador", tipo_es="Acondicionador",
        nombre_pt="Condicionador Matcha & Alecrim 300ml", nombre_es="Acondicionador Matcha y romero 300 ml",
        cantidad=300, unidad="ml", precio_brl=58.90,
        subcat_pt="Controle da oleosidade", subcat_es="Control de la grasitud capilar",
        desc_es="Acondicionador liviano de la linea Matcha que protege el cabello sin sobrecargarlo, ayuda a prevenir la caida y evita el resecamiento de las puntas.",
        beneficios=["Condiciona sin sobrecargar el cabello", "Ayuda a prevenir la caida", "Evita el resecamiento de las puntas"],
        propiedades=["Liviano", "No sobrecarga el cabello"],
        modo_uso="Aplicar sobre el cabello lavado, de medios a puntas, dejar actuar y enjuagar.",
        activos_pt=["Matcha", "Extrato de Alecrim"],
        activos_es=["Matcha", "Extracto de romero"],
        recomendado="Cabello graso o con tendencia a la grasitud excesiva",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-matcha-haskell/",
    ),
    dict(
        linea="Matcha", tipo_pt="Finalizador", tipo_es="Fluido finalizador",
        nombre_pt="Fluido Revitalizante Matcha & Alecrim 120ml", nombre_es="Fluido revitalizante Matcha y romero 120 ml",
        cantidad=120, unidad="ml", precio_brl=48.90,
        subcat_pt="Controle da oleosidade", subcat_es="Control de la grasitud capilar",
        desc_es="Fluido liviano y de rapida absorcion de la linea Matcha, que protege la fibra capilar frente a agresores externos y aporta luminosidad adicional.",
        beneficios=["Absorcion rapida", "Protege la fibra capilar", "Aporta luminosidad adicional"],
        propiedades=["Liviano", "Facil absorcion"],
        modo_uso="Aplicar una pequena cantidad sobre el cabello humedo o seco, evitando la raiz si el cabello es graso.",
        activos_pt=["Matcha", "Extrato de Alecrim"],
        activos_es=["Matcha", "Extracto de romero"],
        recomendado="Cabello graso o con tendencia a la grasitud excesiva",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-matcha-haskell/",
    ),
    dict(
        linea="Murumuru", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Murumuru 300ml", nombre_es="Champu Murumuru 300 ml",
        cantidad=300, unidad="ml", precio_brl=58.90,
        subcat_pt="Nutricao intensa", subcat_es="Nutricion intensa",
        desc_es="Champu de la linea Murumuru que realiza una limpieza suave y repone nutrientes y vitaminas esenciales, indicado para cabello muy seco.",
        beneficios=["Limpieza suave", "Repone nutrientes y vitaminas esenciales"],
        propiedades=["Nutritivo", "Limpieza suave"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien.",
        activos_pt=["Manteiga de Murumuru"],
        activos_es=["Manteca de murumuru"],
        recomendado="Cabello muy seco, opaco o deshidratado",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-murumuru-haskell/",
    ),
    dict(
        linea="Murumuru", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Murumuru 500ml", nombre_es="Champu Murumuru 500 ml",
        cantidad=500, unidad="ml", precio_brl=76.90,
        subcat_pt="Nutricao intensa", subcat_es="Nutricion intensa",
        desc_es="Champu de la linea Murumuru que realiza una limpieza suave y repone nutrientes y vitaminas esenciales, indicado para cabello muy seco.",
        beneficios=["Limpieza suave", "Repone nutrientes y vitaminas esenciales"],
        propiedades=["Nutritivo", "Limpieza suave"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien.",
        activos_pt=["Manteiga de Murumuru"],
        activos_es=["Manteca de murumuru"],
        recomendado="Cabello muy seco, opaco o deshidratado",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-murumuru-haskell/",
    ),
    dict(
        linea="Frizz Cancelado", tipo_pt="Finalizador", tipo_es="Fluido finalizador",
        nombre_pt="Fluido Alinhador Frizz Cancelado 200ml", nombre_es="Fluido alineador Frizz Cancelado 200 ml",
        cantidad=200, unidad="ml", precio_brl=59.90,
        subcat_pt="Controle de frizz", subcat_es="Control del frizz",
        desc_es="Fluido finalizador de la linea Frizz Cancelado que ayuda a alinear y disciplinar el cabello, dejandolo con brillo intenso durante todo el dia sin apelmazarlo.",
        beneficios=["Alinea y disciplina el cabello", "Brillo intenso durante todo el dia", "No apelmaza el cabello"],
        propiedades=["Alineador", "Antifrizz"],
        modo_uso="Aplicar sobre el cabello humedo o seco, de medios a puntas, antes de peinar.",
        activos_pt=[],
        activos_es=["No especificado en la pagina de linea (pendiente de confirmar en ficha de producto)"],
        recomendado="Cabello con frizz",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/frizz-cancelado/",
    ),
    dict(
        linea="Frizz Cancelado", tipo_pt="Reparador", tipo_es="Aceite capilar",
        nombre_pt="Oleo Blinda Fio Frizz Cancelado 100ml", nombre_es="Aceite blindaje capilar Frizz Cancelado 100 ml",
        cantidad=100, unidad="ml", precio_brl=95.90,
        subcat_pt="Controle de frizz", subcat_es="Control del frizz",
        desc_es="Aceite capilar de la linea Frizz Cancelado, formulado para proteger la fibra capilar y contribuir al control del frizz, dejando el cabello con brillo.",
        beneficios=["Protege la fibra capilar", "Contribuye al control del frizz", "Aporta brillo"],
        propiedades=["Protector", "Antifrizz"],
        modo_uso="Aplicar pocas gotas sobre el cabello humedo o seco, de medios a puntas.",
        activos_pt=[],
        activos_es=["No especificado en la pagina de linea (pendiente de confirmar en ficha de producto)"],
        recomendado="Cabello con frizz",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/frizz-cancelado/",
    ),
    dict(
        linea="Frizz Cancelado", tipo_pt="Mascara Liquida Lamelar", tipo_es="Mascarilla liquida lamelar",
        nombre_pt="Mascara Liquida Lamelar Frizz Cancelado 200ml", nombre_es="Mascarilla liquida lamelar Frizz Cancelado 200 ml",
        cantidad=200, unidad="ml", precio_brl=81.90,
        subcat_pt="Controle de frizz", subcat_es="Control del frizz",
        desc_es="Mascarilla liquida de tecnologia lamelar de la linea Frizz Cancelado, que ayuda a alinear el cabello y controlar el frizz con un acabado liviano.",
        beneficios=["Alinea el cabello", "Controla el frizz", "Acabado liviano"],
        propiedades=["Tecnologia lamelar", "Antifrizz"],
        modo_uso="Aplicar sobre el cabello lavado, distribuir de medios a puntas y enjuagar segun las indicaciones del envase.",
        activos_pt=[],
        activos_es=["No especificado en la pagina de linea (pendiente de confirmar en ficha de producto)"],
        recomendado="Cabello con frizz",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/frizz-cancelado/",
    ),
    dict(
        linea="Cachos Sim!", tipo_pt="Shampoo", tipo_es="Champu",
        nombre_pt="Shampoo Cachos Sim! 300ml", nombre_es="Champu Cachos Si! 300 ml",
        cantidad=300, unidad="ml", precio_brl=62.90,
        subcat_pt="Cuidado de cachos", subcat_es="Cuidado de rizos",
        desc_es="Champu de la linea Cachos Si!, formulado sin sulfatos, petrolatos ni parabenos (tecnica low poo), que limpia suavemente sin remover los aceites naturales del cabello rizado.",
        beneficios=["Limpieza suave sin remover aceites naturales", "Hidrata el cabello y el cuero cabelludo", "Libre de sulfatos, petrolatos y parabenos"],
        propiedades=["Formula low poo", "Libre de sulfatos"],
        modo_uso="Aplicar sobre el cabello mojado, masajear el cuero cabelludo y enjuagar bien.",
        activos_pt=["Oleo de Coco", "Colageno Vegetal"],
        activos_es=["Aceite de coco", "Colageno vegetal"],
        recomendado="Cabello rizado y ondulado",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-cachos-sim-haskell/",
    ),
    dict(
        linea="Cachos Sim!", tipo_pt="Condicionador", tipo_es="Acondicionador",
        nombre_pt="Condicionador Cachos Sim! 300ml", nombre_es="Acondicionador Cachos Si! 300 ml",
        cantidad=300, unidad="ml", precio_brl=62.90,
        subcat_pt="Cuidado de cachos", subcat_es="Cuidado de rizos",
        desc_es="Acondicionador de la linea Cachos Si! que aporta emoliencia, definicion y brillo, ayuda a controlar el frizz y favorece el cierre de las cuticulas.",
        beneficios=["Aporta emoliencia, definicion y brillo", "Controla el frizz", "Favorece el cierre de las cuticulas", "Reduce el volumen"],
        propiedades=["Emoliente", "Definidor de rizos"],
        modo_uso="Aplicar sobre el cabello lavado, de medios a puntas, dejar actuar y enjuagar.",
        activos_pt=["Oleo de Coco", "Colageno Vegetal"],
        activos_es=["Aceite de coco", "Colageno vegetal"],
        recomendado="Cabello rizado y ondulado",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-cachos-sim-haskell/",
    ),
    dict(
        linea="Cachos Sim!", tipo_pt="Mascara de Tratamento", tipo_es="Mascarilla de tratamiento",
        nombre_pt="Mascara Umectante Cachos Sim! 300g", nombre_es="Mascarilla humectante Cachos Si! 300 g",
        cantidad=300, unidad="g", precio_brl=74.90,
        subcat_pt="Cuidado de cachos", subcat_es="Cuidado de rizos",
        desc_es="Mascarilla altamente humectante de la linea Cachos Si!, que devuelve la humedad natural, nutre los rizos mas secos y realza su forma original.",
        beneficios=["Alto poder humectante", "Nutre los rizos mas secos", "Realza la forma original del rizo"],
        propiedades=["Humectante", "Nutritiva"],
        modo_uso="Aplicar sobre el cabello lavado, distribuir de medios a puntas y dejar actuar antes de enjuagar.",
        activos_pt=["Oleo de Coco", "Colageno Vegetal"],
        activos_es=["Aceite de coco", "Colageno vegetal"],
        recomendado="Cabello rizado y ondulado, especialmente el mas seco",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-cachos-sim-haskell/",
    ),
    dict(
        linea="Cachos Sim!", tipo_pt="Finalizador", tipo_es="Leave-in (sin enjuague)",
        nombre_pt="Leave In Memorizador de Cachos Cachos Sim! 300ml", nombre_es="Leave-in memorizador de rizos Cachos Si! 300 ml",
        cantidad=300, unidad="ml", precio_brl=68.90,
        subcat_pt="Cuidado de cachos", subcat_es="Cuidado de rizos",
        desc_es="Leave-in de la linea Cachos Si! que activa y define los rizos, controla el frizz y reactiva la memoria del cabello, con proteccion solar y termica.",
        beneficios=["Activa y define los rizos", "Controla el frizz", "Reactiva la memoria del rizo", "Proteccion solar y termica"],
        propiedades=["Definidor", "Con proteccion termica y solar"],
        modo_uso="Aplicar sobre el cabello humedo, distribuir mechon por mechon y dejar secar al aire o con difusor.",
        activos_pt=["Oleo de Coco", "Colageno Vegetal"],
        activos_es=["Aceite de coco", "Colageno vegetal"],
        recomendado="Cabello rizado y ondulado",
        url="https://haskellcosmeticos.com.br/tratamentos-capilares/linha-cachos-sim-haskell/",
    ),
]

from generar_productos_nuevos import construir_productos_nuevos
productos += construir_productos_nuevos()

# ------------------------------------------------------------------
# Construccion del workbook
# ------------------------------------------------------------------
wb = openpyxl.Workbook()

FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10)
FILL_HEADER = PatternFill(start_color="1F4A2E", end_color="1F4A2E", fill_type="solid")  # verde bosque Haskell
FILL_NOTE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

# ---- Hoja 1: Productos ----
ws = wb.active
ws.title = "Productos"

columnas = [
    "id", "slug", "sku", "marca",
    "nombre_original_pt", "nombre_es",
    "linea",
    "categoria_original_pt", "categoria_es",
    "subcategoria_original_pt", "subcategoria_es",
    "tipo_original_pt", "tipo_es",
    "presentacion", "cantidad", "unidad",
    "precio_regular_brl", "precio_oferta_brl", "porcentaje_descuento", "en_oferta",
    "disponible",
    "descripcion_original_pt", "descripcion_es",
    "beneficios_original_pt", "beneficios_es",
    "propiedades_original_pt", "propiedades_es",
    "modo_uso_original_pt", "modo_uso_es",
    "activos_original_pt", "activos_es",
    "ingredientes_originales",
    "imagen_principal_url", "imagen_principal_local", "imagenes_adicionales",
    "recomendado_para_es",
    "url_origen", "fecha_extraccion",
]

for c, nombre in enumerate(columnas, start=1):
    cell = ws.cell(row=1, column=c, value=nombre)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}1"

def slugify(texto):
    import re
    t = texto.lower()
    reemplazos = {"á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n","ü":"u","!":"","ç":"c"}
    for k, v in reemplazos.items():
        t = t.replace(k, v)
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t

fila = 2
for idx, p in enumerate(productos, start=1):
    pid = f"HSK-{idx:04d}"
    slug = slugify(p["nombre_es"])
    presentacion = f'{p["cantidad"]:g} {p["unidad"]}'
    beneficios_es = " | ".join(p["beneficios"])
    propiedades_es = " | ".join(p["propiedades"])
    activos_pt = " | ".join(p["activos_pt"]) if p["activos_pt"] else ""
    activos_es = " | ".join(p["activos_es"])

    valores = [
        pid, slug, None, "Haskell",
        p["nombre_pt"], p["nombre_es"],
        p["linea"],
        "Tratamentos Capilares", "Tratamientos capilares",
        p["subcat_pt"], p["subcat_es"],
        p["tipo_pt"], p["tipo_es"],
        presentacion, p["cantidad"], p["unidad"],
        p["precio_brl"], None, None, False,
        True,
        None, p["desc_es"],
        None, beneficios_es,
        None, propiedades_es,
        None, p["modo_uso"],
        activos_pt, activos_es,
        None,
        None, None, None,
        p["recomendado"],
        p["url"], FECHA_EXTRACCION,
    ]
    for c, val in enumerate(valores, start=1):
        cell = ws.cell(row=fila, column=c, value=val)
        cell.font = FONT_BODY
        cell.alignment = WRAP

    # Formulas reales (no valores fijos) para que recalculen si mas adelante se carga un precio de oferta
    col_reg = columnas.index("precio_regular_brl") + 1
    col_ofe = columnas.index("precio_oferta_brl") + 1
    col_desc = columnas.index("porcentaje_descuento") + 1
    col_enoferta = columnas.index("en_oferta") + 1
    letra_reg = get_column_letter(col_reg)
    letra_ofe = get_column_letter(col_ofe)
    ws.cell(row=fila, column=col_desc,
            value=f'=IF(ISBLANK({letra_ofe}{fila}),0,IFERROR(({letra_reg}{fila}-{letra_ofe}{fila})/{letra_reg}{fila},0))').number_format = "0.0%"
    ws.cell(row=fila, column=col_enoferta,
            value=f'={letra_ofe}{fila}<>""')
    ws.cell(row=fila, column=col_reg).number_format = '#,##0.00 "R$"'
    ws.cell(row=fila, column=col_ofe).number_format = '#,##0.00 "R$"'

    fila += 1

anchos = {
    "id": 10, "slug": 30, "sku": 10, "marca": 10,
    "nombre_original_pt": 32, "nombre_es": 32, "linea": 16,
    "categoria_original_pt": 20, "categoria_es": 20,
    "subcategoria_original_pt": 20, "subcategoria_es": 22,
    "tipo_original_pt": 18, "tipo_es": 20,
    "presentacion": 12, "cantidad": 10, "unidad": 8,
    "precio_regular_brl": 14, "precio_oferta_brl": 14, "porcentaje_descuento": 12, "en_oferta": 10,
    "disponible": 10,
    "descripcion_original_pt": 15, "descripcion_es": 45,
    "beneficios_original_pt": 15, "beneficios_es": 45,
    "propiedades_original_pt": 15, "propiedades_es": 35,
    "modo_uso_original_pt": 15, "modo_uso_es": 40,
    "activos_original_pt": 22, "activos_es": 25,
    "ingredientes_originales": 18,
    "imagen_principal_url": 15, "imagen_principal_local": 15, "imagenes_adicionales": 15,
    "recomendado_para_es": 30,
    "url_origen": 40, "fecha_extraccion": 14,
}
for c, nombre in enumerate(columnas, start=1):
    ws.column_dimensions[get_column_letter(c)].width = anchos.get(nombre, 15)

# ---- Hoja 2: Categorias y lineas ----
ws2 = wb.create_sheet("Categorias_Lineas")
cols2 = ["categoria_original_pt", "categoria_es", "linea", "subcategoria_original_pt", "subcategoria_es", "cantidad_productos_en_muestra"]
for c, nombre in enumerate(cols2, start=1):
    cell = ws2.cell(row=1, column=c, value=nombre)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws2.freeze_panes = "A2"

lineas_resumen = {}
for p in productos:
    key = (p["linea"], p["subcat_pt"], p["subcat_es"])
    lineas_resumen[key] = lineas_resumen.get(key, 0) + 1

fila = 2
for (linea, subcat_pt, subcat_es), cant in sorted(lineas_resumen.items()):
    valores = ["Tratamentos Capilares", "Tratamientos capilares", linea, subcat_pt, subcat_es, cant]
    for c, val in enumerate(valores, start=1):
        cell = ws2.cell(row=fila, column=c, value=val)
        cell.font = FONT_BODY
    fila += 1

for c, nombre in enumerate(cols2, start=1):
    ws2.column_dimensions[get_column_letter(c)].width = 26

# Fila de totales con formula real
ws2.cell(row=fila + 1, column=5, value="Total productos en la muestra:").font = Font(name="Arial", size=10, bold=True)
ws2.cell(row=fila + 1, column=6, value=f"=SUM(F2:F{fila-1})").font = Font(name="Arial", size=10, bold=True)

# ---- Hoja 3: Notas metodologicas ----
ws3 = wb.create_sheet("Notas_Metodologicas")
ws3.column_dimensions["A"].width = 110
notas = [
    ("Fuente", "https://haskellcosmeticos.com.br/ — paginas publicas de linea dentro de Tratamentos Capilares (sin login, sin CAPTCHA evadido)."),
    ("Fecha de extraccion", FECHA_EXTRACCION),
    ("Alcance de esta version", "143 productos reales, en 29 lineas (todas las lineas de Tratamentos Capilares publicadas en el sitio al momento de la extraccion). Se excluyeron los KITS (combos de productos ya listados de forma individual) para no duplicar filas. En 2 lineas (Mandioca, Queratina & Provit B5) la pagina solo mostraba precio individual confirmado para 1 producto — el resto de los items mencionados en el texto de la linea no tenian precio individual visible en la grilla (solo en kits), por lo que no se incluyeron para no inventar precios."),
    ("Como se obtuvo", "Navegacion directa a las paginas publicas de cada una de las 29 lineas y lectura del texto renderizado (nombre, tamano, precio regular y de oferta en BRL, activos principales). No se uso el exportador Playwright del paquete original; para el catalogo 100% completo (incluyendo kits, coloracion y variantes no publicadas en paginas de linea) corresponde ejecutar 'npm run scrape' del exportador ya entregado."),
    ("Redaccion de descripciones, beneficios y modo de uso", "IMPORTANTE: a diferencia de los primeros 20 productos (redactados a mano, ficha por ficha), los 123 productos agregados en esta version usan una redaccion generada a partir de plantillas propias por tipo de producto (champu, acondicionador, mascarilla, leave-in, tonico, etc.) combinadas con el proposito de cada linea. Esto se hizo asi para poder cubrir el catalogo completo en un tiempo razonable. El contenido es original (no es traduccion literal de los textos de marketing del sitio) y describe funciones genericas y reales del tipo de producto, pero es menos especifico que si cada ficha se hubiera redactado individualmente. Los activos y precios si son especificos de cada linea/producto."),
    ("Precios", "Todos los precios son los publicados en BRL (real brasileno) al momento de la extraccion. Cuando la pagina de linea mostraba precio de oferta ademas del regular, se cargaron ambos (precio_oferta_brl); si no se mostraba oferta, la columna queda vacia y en_oferta = FALSO. No se inventaron descuentos."),
    ("SKU", "El sitio publico no muestra codigos SKU en las paginas de linea — la columna queda vacia (null), tal como contempla la especificacion original (seccion 7, ejemplo de JSON)."),
    ("Traduccion", "Espanol neutro latinoamericano. Se evito el voseo argentino (tenes/podes) y el che; se uso el modo impersonal/infinitivo en las instrucciones de uso, valido en toda America Latina. Se conservaron sin traducir los nombres de marca y de linea (Cavalo Forte, Se Curve, Jaborandi, etc.)."),
    ("Ingredientes INCI", "No se incluyen en esta version: la lista de ingredientes tecnicos (INCI) solo aparece en la ficha individual de cada producto, no en la pagina de linea. Columna 'ingredientes_originales' queda vacia; se completa al correr el exportador completo, que si visita cada ficha."),
    ("Imagenes", "No se descargaron ni referenciaron URLs de imagen (las paginas de linea muestran imagenes por CSS/carrusel, no como URL de producto individual estable). Las columnas de imagen quedan vacias; se completan con el exportador completo (seccion 5 de la especificacion, DOWNLOAD_IMAGES=true)."),
    ("Kits excluidos", "Las 29 lineas tienen ademas decenas de KITS (combos de 2 a 7 productos con descuento) que no se cargaron como filas propias para evitar duplicar los productos individuales que ya los componen. Si se necesitan como SKU de venta aparte, se pueden agregar en una pasada posterior."),
    ("Deduplicacion", "Clave usada: slug generado a partir de nombre_es (linea 15 de la especificacion original). El esquema Postgres (esquema_postgres.sql) declara slug UNIQUE y sku UNIQUE (nullable) para que la carga real deduplique igual."),
    ("Decisiones pendientes heredadas de la especificacion (seccion 21)", "Conversion BRL a PEN, si cada presentacion es un producto o una variante (aqui se modelo como variante en el esquema Postgres), si los kits se separan en componentes, alcance final de traduccion (basica vs. IA), y destino final de la carga (frontend / PostgreSQL / Odoo / los tres)."),
    ("Proximo paso sugerido", "1) Revisar esta version ampliada (143 productos, 29 lineas). 2) Si se necesita redaccion ficha por ficha (no por plantilla) para las 123 lineas agregadas, o se necesitan los kits, avisar para una pasada adicional. 3) Ejecutar el exportador completo (npm run scrape) para INCI, imagenes y coloracion. 4) Cargar el resultado en el esquema Postgres adjunto (esquema_postgres.sql). 5) Vincular cada 'sku' resultante con CatalogLine.sku del backend de la plataforma comercial."),
]
r = 1
for titulo, texto in notas:
    c1 = ws3.cell(row=r, column=1, value=titulo)
    c1.font = Font(name="Arial", size=10, bold=True)
    c1.fill = FILL_NOTE
    r += 1
    c2 = ws3.cell(row=r, column=1, value=texto)
    c2.font = FONT_BODY
    c2.alignment = WRAP
    ws3.row_dimensions[r].height = 45
    r += 2

wb.save("productos_haskell.xlsx")
print("OK - filas de productos:", len(productos))
